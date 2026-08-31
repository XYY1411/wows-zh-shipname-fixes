#!/usr/bin/env python3
"""生成 mod 发布文件, 同时提供完整版与增量版。

- 完整版 global.mo:  键集合与该语言官方一致(含空翻译键), 舰船键替换为最终翻译,
                     给『未安装 Localization Loader』的玩家(传统全量覆盖方式)。
- 增量版 wowsZhShipnameFixes.mo: 只含「最终翻译 ≠ 官方 zh/zh_sg 任一」的舰船键,
                     加 X-LocalizationLoader-Priority 头, 给『使用阿斯兰/Localization Loader』的玩家。

每个版本(standard=缩写键 / full=全名版)下, 两种 .mo 都按 zh(国服简中) 与 zh_sg(国际服简中) 输出。

用法:
    python scripts/build_release.py                # 默认最新版本
    python scripts/build_release.py 15.7.0         # 指定版本
    python scripts/build_release.py 15.7.0 --release   # 正式发布: 递增修订号, 不覆盖
"""
import array
import csv
import re
import shutil
import struct
import sys
from datetime import datetime
from pathlib import Path

import polib
from openpyxl import load_workbook

CSV_HEADER = ["键值(key)", "翻译(msgstr)"]
RELEASE_DIR_NAME = "release"  # 发布文件夹名(仓库根目录下)
MOD_NAME = "wowsZhShipnameFixes"  # mod 名(驼峰, 用作增量 .mo 文件名, 不能叫 global.mo)
LOCALES = ("zh", "zh_sg")  # 语言目录: zh=国服简中, zh_sg=国际服简中
LC_PRIORITY = "50"  # LocalizationLoader 优先级(缺省即 50, 并非必须)
# 无后缀舰船键: IDS_P?S???? (不带 _FULL); 对应 _FULL 键为其完整名称
SHIP_BASE_KEY_RE = re.compile(r"^IDS_P[A-Z]S[A-Z]\d{3}$")


def next_release_dir(repo: Path, version: str) -> Path:
    """正式发布: 找 <version>-r<n> 中最大的 n, 返回下一个版本目录 <version>-r(n+1)"""
    base = repo / RELEASE_DIR_NAME
    pattern = re.compile(rf"^{re.escape(version)}-r(\d+)$")
    max_n = 0
    if base.exists():
        for d in base.iterdir():
            if d.is_dir():
                m = pattern.match(d.name)
                if m:
                    max_n = max(max_n, int(m.group(1)))
    n = max_n + 1
    return base / f"{version}-r{n}"


def write_version_txt(release_dir: Path, version: str, mod_version: str, variant: str = "",
                      full_count: int = 0, inc_count: int = 0) -> None:
    """写 mod 版本元数据文件"""
    type_label = "全名版(_FULL)" if variant == "full" else "标准版(缩写键)"
    lines = [
        f"mod版本:      {mod_version}",
        f"对应游戏版本:  {version}",
        f"版本类型:      {type_label}",
        f"语言:          {'/'.join(LOCALES)}",
        f"完整版键数:    {full_count}",
        f"增量键数:      {inc_count}",
        f"构建时间:      {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    ]
    (release_dir / "version.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"  [版本] {release_dir.name}/version.txt (mod {mod_version})")


def find_latest_version(repo: Path) -> str:
    """从 translations/ 找出最大的版本号"""
    versions = []
    for d in (repo / "translations").iterdir():
        if not d.is_dir():
            continue
        parts = d.name.split(".")
        if len(parts) == 3:
            try:
                versions.append((tuple(int(x) for x in parts), d.name))
            except ValueError:
                pass
    if not versions:
        raise SystemExit("translations/ 下没有版本目录")
    versions.sort()
    return versions[-1][1]


def ship_xlsm_to_map(ship_xlsm: Path) -> dict:
    """读取 ship.xlsm 的 (键值, 最终翻译) -> dict"""
    wb = load_workbook(ship_xlsm, read_only=True, data_only=True)
    ws = wb.active
    ship_map = {}
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        key = row[0]
        if key:
            ship_map[str(key)] = str(row[4] or "")
    wb.close()
    print(f"  [读取] ship.xlsm 最终翻译 {len(ship_map)} 条")
    return ship_map


def read_csv_map(csv_path: Path) -> dict:
    """读取两列 CSV(键值, 翻译) -> dict, 键首行表头跳过"""
    data = {}
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        for row in csv.reader(f):
            if row and row[0] and row[0] != CSV_HEADER[0]:
                data[row[0]] = row[1] if len(row) > 1 else ""
    return data


def effective_translations(ship_map: dict, variant: str) -> dict:
    """得到该版本应生效的翻译表:
    standard: 直接用最终翻译(缩写键);
    full:     无后缀键 IDS_P?S???? 用对应 _FULL 键的完整船名。
    """
    eff = dict(ship_map)
    if variant == "full":
        for k in list(eff):
            if SHIP_BASE_KEY_RE.match(k):
                full_k = k + "_FULL"
                if full_k in ship_map:
                    eff[k] = ship_map[full_k]
    return eff


def po_to_mo(po: polib.POFile, mo_path: Path) -> None:
    """自定义 mo 编译器: 遍历全部条目(含空翻译键), 保证键集合与官方一致。
    不用 polib.save_as_mofile(其 translated_entries() 会丢空翻译键, 导致游戏显示 IDS_XXX)。
    复数条目(msgid_plural)按 gettext 规范用 NUL 连接写入。
    """
    entries = list(po)
    entries.sort(key=lambda o: o.msgid_with_context.encode("utf-8"))
    entries = [po.metadata_as_entry()] + entries
    entries_len = len(entries)
    ids, strs = b"", b""
    offsets = []
    for e in entries:
        msgid = b""
        if e.msgctxt:
            msgid = po._encode(e.msgctxt + "\4")
        if e.msgid_plural:
            msgstr = []
            for index in sorted(e.msgstr_plural.keys()):
                msgstr.append(e.msgstr_plural[index])
            msgid += po._encode(e.msgid + "\0" + e.msgid_plural)
            msgstr = po._encode("\0".join(msgstr))
        else:
            msgid += po._encode(e.msgid)
            msgstr = po._encode(e.msgstr)
        offsets.append((len(ids), len(msgid), len(strs), len(msgstr)))
        ids += msgid + b"\0"
        strs += msgstr + b"\0"
    keystart = 7 * 4 + 16 * entries_len
    valuestart = keystart + len(ids)
    koffsets, voffsets = [], []
    for o1, l1, o2, l2 in offsets:
        koffsets += [l1, o1 + keystart]
        voffsets += [l2, o2 + valuestart]
    offsets = koffsets + voffsets
    output = struct.pack(
        "Iiiiiii",
        0x950412DE,  # gettext magic
        0, entries_len, 7 * 4, 7 * 4 + entries_len * 8, 0, keystart
    )
    output += array.array("i", offsets).tobytes()
    output += ids
    output += strs
    mo_path.write_bytes(output)
    print(f"  [完整] {mo_path} ({entries_len - 1} 条, 含空翻译键)")


def build_full_mo(official_map: dict, ship_map: dict, variant: str, out_mo: Path) -> int:
    """生成完整版 global.mo: 该语言官方全部键 + 舰船键覆盖为最终翻译, 返回总键数"""
    eff = effective_translations(ship_map, variant)
    merged = dict(official_map)
    for k, v in eff.items():
        if v:  # 只覆盖有最终翻译的舰船键
            merged[k] = v
    po = polib.POFile()
    po.metadata = {
        "Project-Id-Version": MOD_NAME,
        "Language": "zh_CN",
        "MIME-Version": "1.0",
        "Content-Type": "text/plain; charset=UTF-8",
        "Content-Transfer-Encoding": "8bit",
    }
    for k in sorted(merged):
        po.append(polib.POEntry(msgid=k, msgstr=merged[k]))
    out_mo.parent.mkdir(parents=True, exist_ok=True)
    po_to_mo(po, out_mo)
    return len(merged)


def build_incremental_mo(ship_map: dict, official_zh: dict, official_zhsg: dict, variant: str, out_mo: Path) -> int:
    """生成增量版 .mo(只含差异键, 给 Localization Loader), 返回差异键数。
    差异判定: 最终翻译非空 且 与 zh(国服) 或 zh_sg(国际服) 任一官方不同才计入。
    """
    eff = effective_translations(ship_map, variant)
    delta = {k: v for k, v in eff.items() if v and (official_zh.get(k) != v or official_zhsg.get(k) != v)}
    if not delta:
        print(f"  [{variant}] 无差异键, 跳过增量")
        return 0

    po = polib.POFile()
    po.metadata = {
        "Project-Id-Version": MOD_NAME,
        "Language": "zh_CN",
        "MIME-Version": "1.0",
        "Content-Type": "text/plain; charset=UTF-8",
        "Content-Transfer-Encoding": "8bit",
        # LocalizationLoader 优先级
        "X-LocalizationLoader-Priority": LC_PRIORITY,
    }
    for k in sorted(delta):
        po.append(polib.POEntry(msgid=k, msgstr=delta[k]))

    out_mo.parent.mkdir(parents=True, exist_ok=True)
    po.save_as_mofile(str(out_mo))  # 增量只含差异键, 无空键; metadata(含优先级头)会写入
    print(f"  [{variant}] 增量 {out_mo} ({len(delta)} 条差异, 优先级 {LC_PRIORITY})")
    return len(delta)


def cleanup_legacy(variant_dir: Path) -> None:
    """删除旧版中间产物(global.csv / global.po, 现在直接用 dict 生成不再输出)"""
    for name in ("global.csv", "global.po"):
        p = variant_dir / name
        if p.exists():
            p.unlink()


def main():
    repo = Path(__file__).resolve().parent.parent
    args = [a for a in sys.argv[1:] if a != "--release"]
    release_mode = "--release" in sys.argv   # 正式发布: 递增修订号, 不覆盖
    version = args[0] if args else find_latest_version(repo)
    print(f"构建 mod 发布版本: {version}")

    if release_mode:
        release_dir = next_release_dir(repo, version)
        mod_version = release_dir.name
        print(f"[正式发布] 生成新版本目录 {release_dir.name} (不覆盖)")
    else:
        release_dir = repo / RELEASE_DIR_NAME / version
        mod_version = version
        print(f"[本地调试] 覆盖最新目录 {release_dir.name}")

    ver_dir = repo / "translations" / version
    if not ver_dir.exists():
        raise SystemExit(f"版本目录不存在: {ver_dir}")
    ship_xlsm = ver_dir / "ship.xlsm"
    if not ship_xlsm.exists():
        raise SystemExit(f"ship.xlsm 不存在: {ship_xlsm}")
    for loc in LOCALES:
        if not (ver_dir / loc / "global.csv").exists():
            raise SystemExit(f"{loc}/global.csv 不存在: {ver_dir / loc / 'global.csv'}")

    ship_map = ship_xlsm_to_map(ship_xlsm)
    official_zh = read_csv_map(ver_dir / "zh" / "global.csv")
    official_zhsg = read_csv_map(ver_dir / "zh_sg" / "global.csv")

    release_dir.mkdir(parents=True, exist_ok=True)
    for variant in ("standard", "full"):
        variant_dir = release_dir / variant
        variant_dir.mkdir(parents=True, exist_ok=True)

        # 1) 完整版 global.mo: 每个语言一份(给无 loader 用户)
        full_count = 0
        for lang, official in (("zh", official_zh), ("zh_sg", official_zhsg)):
            full_mo = variant_dir / lang / "LC_MESSAGES" / "global.mo"
            full_count = build_full_mo(official, ship_map, variant, full_mo)

        # 2) 增量版 wowsZhShipnameFixes.mo: 给 loader 用户(zh 生成后复制到 zh_sg)
        zh_mo = variant_dir / "zh" / "LC_MESSAGES" / f"{MOD_NAME}.mo"
        inc_count = build_incremental_mo(ship_map, official_zh, official_zhsg, variant, zh_mo)
        if inc_count:
            zhsg_mo = variant_dir / "zh_sg" / "LC_MESSAGES" / f"{MOD_NAME}.mo"
            zhsg_mo.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(zh_mo, zhsg_mo)
            print(f"  [{variant}] 复制到 {zhsg_mo}")

        write_version_txt(variant_dir, version,
                          mod_version if variant == "standard" else f"{mod_version}-full",
                          variant, full_count=full_count, inc_count=inc_count)
        cleanup_legacy(variant_dir)  # 删除旧 global.csv/po 中间产物

    print(f"\n完成! 发布文件位于: {release_dir}/{{standard,full}}/{{zh,zh_sg}}/LC_MESSAGES/"
          f"{{global.mo, {MOD_NAME}.mo}}")


if __name__ == "__main__":
    sys.exit(main())
