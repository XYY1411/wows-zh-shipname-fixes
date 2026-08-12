#!/usr/bin/env python3
"""由各版本的三语言 CSV 生成:
  - global.xlsx         三语言对照表(不着色)
  - ship.xlsm           舰船词条 VB 表格(基于 templates/ship_template.xlsm)
  - global_diff.xlsx    与上一版本的差异(全部词条)
  - ship_diff.xlsx      与上一版本的差异(舰船词条, 无宏)

用法:
  python scripts/generate_tables.py             # 自动取最新两个版本
  python scripts/generate_tables.py <新版本> <旧版本>
"""
import csv
import re
import shutil
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

LANGS = ["zh", "zh_sg", "en"]
HEADERS = ["键值(key)", "简体中文(zh)", "简体中文新加坡(zh_sg)", "英文(en)", "最终翻译"]
DIFF_HEADERS = ["键值(key)", "差异类型",
                "简体中文(旧)", "简体中文(新)", "简体中文新加坡(旧)", "简体中文新加坡(新)",
                "英文(旧)", "英文(新)", "最终翻译"]
SHIP_KEY_RE = re.compile(r"^IDS_P[A-Z]S[A-Z]\d{3}(?:_FULL)?$")
# polib 解析时生成的 po 元数据伪条目, 不是真实词条, 需要过滤
META_KEYS = {"#PluralForms", ""}
DIFF_TYPES = {"added": "新增", "removed": "删除", "modified": "修改"}
HEADER_FILL = "FF4472C4"  # 模板表头同款蓝色
DATA_FONT = Font(name="Consolas", size=11)  # 数据区等宽字体, 系统自带无需安装


def apply_data_font(ws, ncols: int):
    """数据区统一使用 Consolas 等宽字体"""
    if ws.max_row < 2:
        return
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ncols):
        for cell in row:
            cell.font = DATA_FONT


def purge_ghost_rows(ws, expected: int) -> int:
    """清理 expected 行之后的幽灵空行(有样式无内容), 返回清理数量"""
    if ws.max_row <= expected:
        return 0
    ghost = ws.max_row - expected
    ws.delete_rows(expected + 1, ghost)
    if ws.max_row > expected:  # 兜底: 清空残留单元格
        for row in ws.iter_rows(min_row=expected + 1):
            for cell in row:
                cell.value = None
    return ghost


def verify_xml_clean(path: Path, expected_rows: int):
    """生成后终极验证: 全面检查远端残留(幽灵行源头)
    1. row 元素最大行号(覆盖所有列)
    2. 单元格引用最大行号(任意列, 如 E 列远端)
    3. 条件格式/数据验证 sqref 引用范围
    4. workbook 打印区域 definedNames 引用范围
    任何一项超出预期数据行都会报错, 防止 Excel 打开后 UsedRange 被撑大。
    """
    import zipfile
    import re as _re
    with zipfile.ZipFile(path) as z:
        sheet = [n for n in z.namelist() if _re.match(r"xl/worksheets/sheet\d+\.xml", n)]
        if not sheet:
            return
        xml = z.read(sheet[0]).decode("utf-8")
        wbxml = z.read("xl/workbook.xml").decode("utf-8") if "xl/workbook.xml" in z.namelist() else ""

    problems = []

    # 1. row 元素最大行号(任何列的行都会产生 row 元素)
    rows = [int(r) for r in _re.findall(r'<row r="(\d+)"', xml)]
    if rows and max(rows) > expected_rows:
        problems.append(f"row 元素最大行号 {max(rows)} > {expected_rows}")

    # 2. 单元格引用最大行号(逐列检查, 任意列远端都逃不掉)
    cells = [int(r) for r in _re.findall(r'<c r="[A-Z]{1,3}(\d+)"', xml)]
    if cells and max(cells) > expected_rows:
        problems.append(f"单元格最大行号 {max(cells)} > {expected_rows}")

    # 3. 条件格式 / 数据验证 sqref 引用范围
    for sq in _re.findall(r'sqref="([^"]+)"', xml):
        for ref in sq.split():
            m = _re.match(r"^[A-Z]{1,3}(\d+)(?::[A-Z]{1,3}(\d+))?$", ref)
            if m:
                r1, r2 = int(m.group(1)), int(m.group(2) or m.group(1))
                if max(r1, r2) > expected_rows:
                    problems.append(f"sqref 引用远端 {ref}")

    # 4. 打印区域 / 其他 definedNames 引用行号
    for dn in _re.findall(r"<definedName[^>]*>([^<]*)</definedName>", wbxml):
        for rr in _re.findall(r"\$?[A-Z]{1,3}\$?(\d+)", dn):
            if int(rr) > expected_rows:
                problems.append(f"definedName 引用远端行 {rr}")

    if problems:
        raise RuntimeError(
            f"幽灵行检测失败 {path.name}: " + "; ".join(problems) +
            "。远端残留会导致 Excel 打开后 UsedRange 撑大, 请检查模板或生成逻辑。"
        )


def read_csv(path: Path) -> dict:
    """读取 global.csv / ship.csv -> {msgid: msgstr}"""
    data = {}
    with open(path, encoding="utf-8-sig", newline="") as f:
        r = csv.reader(f)
        next(r, None)
        for row in r:
            if len(row) >= 2 and row[0] not in META_KEYS:
                data[row[0]] = row[1]
    return data


def merge_langs(version_dir: Path) -> dict:
    """合并三语言 -> {key: {zh, zh_sg, en}}"""
    merged = {}
    for lang in LANGS:
        d = read_csv(version_dir / lang / "global.csv")
        for k, v in d.items():
            merged.setdefault(k, {})[lang] = v
    return merged


def style_header(ws, ncols: int):
    """表头: Arial 粗体白字 + 蓝色填充(与模板一致)"""
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = Font(name="Arial", bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)


def make_global_xlsx(version_dir: Path, data: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "翻译对照"
    ws.append(HEADERS)
    style_header(ws, 5)
    for k in sorted(data):
        d = data[k]
        ws.append([k, d.get("zh", ""), d.get("zh_sg", ""), d.get("en", ""), ""])
    for col, w in zip("ABCDE", [26, 50, 50, 50, 50]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    apply_data_font(ws, 5)
    ghost = purge_ghost_rows(ws, len(data) + 1)
    out = version_dir / "global.xlsx"
    wb.save(out)
    verify_xml_clean(out, len(data) + 1)
    print(f"  [生成] {out} ({ws.max_row - 1} 行, 幽灵行清理 {ghost})")


def read_translations(ship_path: Path) -> dict:
    """读取 ship.xlsm 的 {键值: 最终翻译} (E 列)"""
    trans = {}
    try:
        wb = load_workbook(ship_path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            key = row[0]
            if key and len(row) >= 5 and row[4]:
                trans[str(key)] = str(row[4]).strip()
        wb.close()
    except Exception as e:
        print(f"  [警告] 读取旧翻译失败 {ship_path}: {e}")
    return trans


def make_ship_xlsm(version_dir: Path, data: dict, template: Path, carry_from: Path = None):
    """复制模板 xlsm(保留 VB 宏与表头样式), 填入舰船词条。
    carry_from: 旧版本 ship.xlsm 路径, 按键值迁移最终翻译(E 列)到新版本。
    新增行(旧版本没有的键)的最终翻译留空。
    """
    ship_data = {k: d for k, d in data.items() if SHIP_KEY_RE.match(k)}
    dst = version_dir / "ship.xlsm"
    trans = {}
    if dst.exists():  # 目标文件自身翻译优先(防止重建时丢失人工翻译)
        trans.update(read_translations(dst))
        print(f"  [保留] 从 {version_dir.name}/ship.xlsm 读取现有翻译 {len(trans)} 条")
    if carry_from is not None and carry_from.exists():  # 旧版本补漏, 不覆盖已有
        for k, v in read_translations(carry_from).items():
            trans.setdefault(k, v)
        print(f"  [迁移] 从 {carry_from.parent.name}/ship.xlsm 补齐翻译")
    shutil.copy2(template, dst)
    wb = load_workbook(dst, keep_vba=True)
    ws = wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    carried = 0
    for k in sorted(ship_data):
        d = ship_data[k]
        old_tr = trans.get(k, "")
        if old_tr:
            carried += 1
        ws.append([k, d.get("zh", ""), d.get("zh_sg", ""), d.get("en", ""), old_tr])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    apply_data_font(ws, 5)
    ghost = purge_ghost_rows(ws, len(ship_data) + 1)
    wb.save(dst)
    verify_xml_clean(dst, len(ship_data) + 1)
    print(f"  [生成] {dst} ({ws.max_row - 1} 行舰船词条, 迁移翻译 {carried} 条, 幽灵行清理 {ghost})")


def diff_data(old: dict, new: dict) -> list:
    """返回差异行: [key, 类型, 旧zh, 新zh, 旧zh_sg, 新zh_sg, 旧en, 新en, 最终翻译]"""
    rows = []
    for k in sorted(set(old) | set(new)):
        if k not in old:
            nd = new[k]
            rows.append([k, DIFF_TYPES["added"], "", nd.get("zh", ""), "",
                         nd.get("zh_sg", ""), "", nd.get("en", ""), ""])
        elif k not in new:
            od = old[k]
            rows.append([k, DIFF_TYPES["removed"], od.get("zh", ""), "",
                         od.get("zh_sg", ""), "", od.get("en", ""), ""])
        else:
            od, nd = old[k], new[k]
            if any(od.get(l, "") != nd.get(l, "") for l in LANGS):
                rows.append([k, DIFF_TYPES["modified"], od.get("zh", ""), nd.get("zh", ""),
                             od.get("zh_sg", ""), nd.get("zh_sg", ""),
                             od.get("en", ""), nd.get("en", ""), ""])
    return rows


def make_diff_xlsx(new_dir: Path, old: dict, new: dict):
    """全部词条差异 -> global_diff.xlsx"""
    rows = diff_data(old, new)
    wb = Workbook()
    ws = wb.active
    ws.title = "翻译差异"
    ws.append(DIFF_HEADERS)
    style_header(ws, len(DIFF_HEADERS))
    for r in rows:
        ws.append(r)
    for col, w in zip("ABCDEFGHI", [26, 10, 45, 45, 45, 45, 45, 45, 20]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"
    apply_data_font(ws, 9)
    ghost = purge_ghost_rows(ws, len(rows) + 1)
    out = new_dir / "global_diff.xlsx"
    wb.save(out)
    verify_xml_clean(out, len(rows) + 1)
    print(f"  [生成] {out} ({len(rows)} 行差异, 幽灵行清理 {ghost})")


def make_ship_diff_xlsx(new_dir: Path, old: dict, new: dict):
    """舰船词条差异 -> ship_diff.xlsx (无宏, 不对比最终翻译列)"""
    rows = [r for r in diff_data(old, new) if SHIP_KEY_RE.match(r[0])]
    wb = Workbook()
    ws = wb.active
    ws.title = "舰船差异"
    ws.append(DIFF_HEADERS)
    style_header(ws, len(DIFF_HEADERS))
    for r in rows:
        ws.append(r)
    for col, w in zip("ABCDEFGHI", [26, 10, 45, 45, 45, 45, 45, 45, 20]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"
    apply_data_font(ws, 9)
    ghost = purge_ghost_rows(ws, len(rows) + 1)
    out = new_dir / "ship_diff.xlsx"
    wb.save(out)
    verify_xml_clean(out, len(rows) + 1)
    print(f"  [生成] {out} ({len(rows)} 行舰船差异, 幽灵行清理 {ghost})")


def find_latest_versions(repo: Path) -> tuple:
    """自动找 translations/ 下最新的两个版本目录"""
    dirs = [d.name for d in (repo / "translations").iterdir()
            if d.is_dir() and re.match(r"^\d+\.\d+\.\d+$", d.name)]
    dirs.sort(key=lambda v: tuple(int(x) for x in v.split(".")))
    if len(dirs) < 2:
        raise SystemExit(f"translations/ 下版本不足两个: {dirs}")
    return dirs[-1], dirs[-2]


def main():
    repo = Path(__file__).resolve().parent.parent
    template = repo / "templates" / "ship_template.xlsm"
    force = "--force" in sys.argv
    args = [a for a in sys.argv[1:] if a != "--force"]

    if len(args) >= 2:
        new_ver, old_ver = args[0], args[1]
    else:
        new_ver, old_ver = find_latest_versions(repo)
        print(f"自动检测版本: 新={new_ver} 旧={old_ver}")

    new_dir = repo / "translations" / new_ver
    old_dir = repo / "translations" / old_ver
    if not new_dir.exists() or not old_dir.exists():
        raise SystemExit(f"版本目录不存在: {new_dir} / {old_dir}")

    print(f"合并 {new_ver} 三语言 CSV ...")
    new_data = merge_langs(new_dir)
    print(f"合并 {old_ver} 三语言 CSV ...")
    old_data = merge_langs(old_dir)

    def want(path: Path) -> bool:
        """增量生成: 文件已存在则跳过(保护人工修改), --force 强制重建"""
        if path.exists() and not force:
            print(f"  [跳过] {path.name} 已存在, 保留不动 (强制重建加 --force)")
            return False
        return True

    if template.exists():
        print(f"检查 {new_ver} 的表格:")
        if want(new_dir / "global.xlsx"):
            make_global_xlsx(new_dir, new_data)
        if want(new_dir / "ship.xlsm"):
            make_ship_xlsm(new_dir, new_data, template, carry_from=old_dir / "ship.xlsm")
        print(f"检查 {old_ver} 的表格:")
        if want(old_dir / "global.xlsx"):
            make_global_xlsx(old_dir, old_data)
        if want(old_dir / "ship.xlsm"):
            make_ship_xlsm(old_dir, old_data, template)
    else:
        print(f"[警告] 模板不存在 {template}, 只生成 xlsx 表格")
        print(f"检查 {new_ver} 的表格:")
        if want(new_dir / "global.xlsx"):
            make_global_xlsx(new_dir, new_data)
        print(f"检查 {old_ver} 的表格:")
        if want(old_dir / "global.xlsx"):
            make_global_xlsx(old_dir, old_data)

    print(f"检查 {new_ver} 相对 {old_ver} 的差异:")
    if want(new_dir / "global_diff.xlsx"):
        make_diff_xlsx(new_dir, old_data, new_data)
    if want(new_dir / "ship_diff.xlsx"):
        make_ship_diff_xlsx(new_dir, old_data, new_data)
    return 0


if __name__ == "__main__":
    sys.exit(main())
